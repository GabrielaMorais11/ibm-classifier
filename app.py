from flask import Flask, request, send_file, jsonify
from flask_cors import CORS
import openpyxl
from io import BytesIO
from datetime import datetime
import os
import unicodedata
import re

app = Flask(__name__, static_folder='.', static_url_path='')
CORS(app)

# Base de dados mestre com TODAS as contas
MASTER_DATA = {
    "AB MAURI": "Select T",
    "ACHE": "Select T",
    "ADAMA": "Select T",
    "AGREX": "Select T",
    "ALL4LABELS": "Select T",
    "BAUMGARTEN GRAFICA": "Select T",
    "ALPARGATAS": "Select T",
    "ANGELONI": "Select T",
    "APERAM SOUTH AMERICA": "Select T",
    "ARCELORMITTAL BRASIL": "Select Horizon",
    "AREZZO CO": "Select T",
    "ATACADAO DIA A DIA": "Select T",
    "AURORA ALIMENTOS": "Select T",
    "AZUL LINHAS AEREAS": "Select T",
    "B3": "Select Horizon",
    "BANCO BS2": "Select T",
    "BONSUCESSO": "Select T",
    "BANCO BV": "Select Horizon",
    "BANCO PINE": "Select T",
    "BANCO RENDIMENTO": "Select T",
    "BAYER": "Select T",
    "BISTEK SUPERMERCADOS": "Select T",
    "BMB": "Select T",
    "CARAMURU": "Select T",
    "CASAS PERNAMBUCANAS": "Select T",
    "CASTROLANDA": "Select T",
    "CEG": "Select T",
    "NATURGY": "Select T",
    "CENIBRA": "Select T",
    "CEREAL COMERCIO": "Select T",
    "CISER": "Select T",
    "CLARO": "Strategic",
    "CMPC": "Select T",
    "COAMO": "Select T",
    "COASUL": "Select T",
    "COCAMAR": "Select T",
    "COELHO DA FONSECA": "Select T",
    "COFCO BRASIL": "Select T",
    "COGNA EDUCACAO": "Select T",
    "CONTINENTAL AUTOMOTIVA DO BRASIL": "Select T",
    "COOPAVEL": "Select T",
    "COOPERCITRUS": "Select T",
    "COPACOL": "Select T",
    "COPA ENERGIA": "Select T",
    "COPASUL": "Select T",
    "COPERSUCAR": "Select T",
    "COPOBRAS": "Select T",
    "COTRIJAL": "Select T",
    "CREDSYSTEM": "Select T",
    "DANNEMANN SIEMSEN BIGLER": "Select T",
    "DAVITA": "Select T",
    "DELL": "Select T",
    "DELLA VIA PNEUS": "Select T",
    "DIGIO": "Enterprise",
    "DIRECIONAL ENGENHARIA": "Select T",
    "DMA": "Select T",
    "ELDORADO BRASIL": "Select Horizon",
    "ELECTROLUX": "Select T",
    "ELFA MEDICAMENTOS": "Select T",
    "ELGIN": "Select T",
    "EMBRATEL STAR ONE": "Strategic",
    "EMBRATEL": "Strategic",
    "STAR ONE": "Strategic",
    "EUCATEX": "Select T",
    "FARMACIA PAGUE MENOS": "Select T",
    "PAGUE MENOS": "Select T",
    "FAST SHOP": "Select T",
    "FAURECIA": "Select T",
    "FIRJAN": "Select Horizon",
    "FRUKI": "Select T",
    "FUNDACAO BUTANTAN": "Select T",
    "GRANOL AGRICOLA": "Select T",
    "GRENDENE": "Select T",
    "GRUPO ARGENTA": "Select T",
    "ARGENTA": "Select T",
    "GRUPO JBS": "Select T",
    "JBS": "Select T",
    "GRUPO J MALUCELLI": "Select T",
    "J MALUCELLI": "Select T",
    "GRUPO MARQUISE": "Select T",
    "MARQUISE": "Select T",
    "GRUPO MOSAIC": "Select T",
    "MOSAIC": "Select T",
    "GRUPO OXXO": "Select T",
    "OXXO": "Select T",
    "GRUPO PARTAGE": "Select T",
    "PARTAGE": "Select T",
    "GRUPO PEREIRA": "Select T",
    "PEREIRA": "Select T",
    "GRUPO SER EDUCACIONAL": "Select T",
    "SER EDUCACIONAL": "Select T",
    "GRUPO TRES CORACOES ALIMENTOS": "Select T",
    "TRES CORACOES": "Select T",
    "GRUPO ULTRA BRASIL": "Select T",
    "ULTRA BRASIL": "Select T",
    "GSH CORP": "Select T",
    "HAPVIDA NOTREDAME INTERMEDICA": "Select T",
    "HAPVIDA": "Select T",
    "NOTREDAME INTERMEDICA": "Select T",
    "HINODE": "Select T",
    "INBRANDS": "Select T",
    "J MACEDO": "Select T",
    "KRONA": "Select T",
    "LARCO PETROLEO": "Select T",
    "LOUIS DREYFUS": "Select T",
    "LIBRELATO": "Select T",
    "LINDT SPRUNGLI": "Select T",
    "LOREAL": "Select T",
    "MACKENZIE": "Select T",
    "MAHLE": "Select T",
    "MARTIN BROWER": "Select T",
    "MART MINAS": "Select T",
    "M CASSAB": "Select T",
    "M DIAS BRANCO": "Select T",
    "MIDEA CARRIER": "Select T",
    "MK ELETRODOMESTICOS": "Select T",
    "MOBILIZE FINANCIAL SERVICES": "Select T",
    "MULTILASER": "Select T",
    "NEOENERGIA ELEKTRO": "Select T",
    "NEUGEBAUER": "Select T",
    "NORSK HYDRO": "Select T",
    "NOVELIS": "Select T",
    "ODONTOPREV": "Enterprise",
    "OLEOPLAN": "Select T",
    "PATRIA INVESTIMENTOS": "Select T",
    "PATRUS TRANSPORTES": "Select T",
    "PETZ": "Select T",
    "PIF PAF": "Select T",
    "PIRELLI PNEUS": "Select T",
    "PORTO BANK": "Select Horizon",
    "POSITIVO TECNOLOGIA": "Select T",
    "PROFARMA": "Select T",
    "PRUMO LOGISTICA": "Select T",
    "PORTO DO ACU": "Select T",
    "RANDON": "Select T",
    "REDE GLOBO": "Select T",
    "RENAULT": "Select T",
    "ROCHE": "Select T",
    "ROCKWELL AUTOMATION": "Select T",
    "RODOIL": "Select T",
    "SAGA BRASIL": "Select T",
    "SAO SALVADOR ALIMENTOS": "Select T",
    "SSA": "Select T",
    "SAP": "Select T",
    "SEPHORA": "Select T",
    "SIDERURGICA USIMINAS": "Select Horizon",
    "USIMINAS": "Select Horizon",
    "SIEMENS": "Select T",
    "SLC AGRICOLA": "Select T",
    "SUPERMERCADOS MUNDIAL": "Select T",
    "SUPERMIX": "Select T",
    "SYLVAMO": "Select T",
    "TENDA ATACADO": "Select T",
    "TERMOMECANICA": "Select T",
    "TIMBRO": "Select T",
    "TITAN PNEUS": "Select T",
    "TK ELEVADORES BRASIL": "Select T",
    "TOKIO MARINE SEGURADORA": "Select T",
    "TT WORK": "Select T",
    "UISA BIOENERGIA": "Select T",
    "UNIPAR CARBOCLORO": "Select T",
    "USINA CORURIPE": "Select T",
    "VERO": "Select T",
    "VIBRA": "Select Horizon",
    "VICUNHA TEXTIL": "Select T",
    "VIPAL": "Select T",
    "VITRU EDUCACAO": "Select T",
    "YARA": "Select T",
    "YPE": "Select T",
    "ZAMP": "Select T",
}

def normalize(text):
    """Remove acentos e normaliza texto"""
    if not text:
        return ""
    text = str(text).upper().strip()
    text = unicodedata.normalize('NFD', text)
    text = ''.join(c for c in text if unicodedata.category(c) != 'Mn')
    text = re.sub(r'[^A-Z0-9\s]', ' ', text)
    text = re.sub(r'\s+', ' ', text).strip()
    return text

def find_classification(account_name):
    """Busca classificação para uma conta"""
    if not account_name:
        return "Não Classificado", "N/A", "0%"
    
    normalized_account = normalize(account_name)
    
    # Busca exata
    for master_key, classification in MASTER_DATA.items():
        if normalize(master_key) == normalized_account:
            return classification, "Exata", "100%"
    
    # Busca parcial
    for master_key, classification in MASTER_DATA.items():
        normalized_master = normalize(master_key)
        if normalized_master in normalized_account or normalized_account in normalized_master:
            return classification, "Parcial", "90%"
    
    return "Não Classificado", "Não encontrado", "0%"

@app.route('/')
def index():
    return app.send_static_file('index.html')

@app.route('/health')
def health():
    return jsonify({
        'status': 'ok',
        'total_contas': len(MASTER_DATA),
        'timestamp': datetime.now().isoformat()
    })

@app.route('/classify', methods=['POST'])
def classify():
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'Nenhum arquivo enviado'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'Nome de arquivo vazio'}), 400
        
        # Ler Excel
        wb = openpyxl.load_workbook(file)
        ws = wb.active
        
        # Adicionar cabeçalhos de classificação
        max_col = ws.max_column
        ws.cell(1, max_col + 1, "Classificacao")
        ws.cell(1, max_col + 2, "Metodo")
        ws.cell(1, max_col + 3, "Confianca")
        
        # Processar cada linha
        for row in range(2, ws.max_row + 1):
            account_name = ws.cell(row, 1).value
            classification, method, confidence = find_classification(account_name)
            
            ws.cell(row, max_col + 1, classification)
            ws.cell(row, max_col + 2, method)
            ws.cell(row, max_col + 3, confidence)
        
        # Salvar resultado
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'contas_classificadas_{timestamp}.xlsx'
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        return jsonify({'error': f'Erro: {str(e)}'}), 500

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 10000))
    app.run(host='0.0.0.0', port=port, debug=False)
